"""수강 이력 서비스 — '전체 기이수성적' 엑셀 업로드 · 수동 추가/삭제 · 평점평균 산출.

엑셀 형식 (학교 포털에서 받는 그대로)
    년도 | 학기 | 과목번호 | 분반 | 과목명 | 이수구분 | 학점 | 등급 | 평점 | 재수강(대체)과목

실측에서 확인한 특성 — 파서가 반드시 처리해야 한다
    · 모든 셀이 문자열이고 평점에는 앞 공백이 붙는다(' 4.5')
    · P(Pass) 과목은 평점이 비어 있다 → 평점평균 계산에서 빼야 한다
    · '여름학기'가 존재하고, 합계 행은 없다

병합(merge) 정책
    같은 학생의 (과목번호, 년도, 학기)가 이미 있으면 그 행을 그대로 두고 건너뛴다.
    재수강은 학기가 달라 별도 행으로 남고, 같은 파일을 다시 올려도 중복이 생기지 않는다.
    (student_course에 DB 유니크 제약을 걸지 않은 이유: 기존 더미 행은 year/semester가
     NULL이라 제약을 추가하면 기존 데이터와 충돌한다. 응용 계층에서 판정한다.)
"""
import io
from collections import defaultdict

from sqlalchemy import select, delete, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.DB_Table import Course, Student, StudentCourse

# 엑셀 이수구분 → course.category.
# 졸업요건 집계(graduation._check_graduation_status)가 이 값들을 그대로 더한다:
#   전공 = 전공필수 + 전공선택 / 교양 = 교양필수 + 교양선택 / 일반 = 일반 / 트랙 = 트랙
# '전공'·'교양'처럼 필수·선택 구분이 없는 값은 선택으로 보낸다 — 합계가 같아 안전하다.
CATEGORY_MAP = {
    "전공": "전공선택", "전공선택": "전공선택", "전공필수": "전공필수",
    "교양": "교양선택", "교양선택": "교양선택", "교양필수": "교양필수",
    "일반": "일반", "일반선택": "일반", "일선": "일반",
    "트랙": "트랙", "다전공": "트랙", "복수전공": "트랙", "부전공": "트랙",
}
DEFAULT_CATEGORY = "일반"

# 미이수 등급 — 학점으로 인정하지 않는다
FAIL_GRADES = {"F", "NP", "U", "W", "I"}

EXPECTED_HEADER = ["년도", "학기", "과목번호", "분반", "과목명", "이수구분", "학점", "등급", "평점"]


def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v) -> float | None:
    t = _s(v).replace(",", "")
    if not t:
        return None
    try:
        return float(t)
    except ValueError:
        return None


class StudentCourseService:

    # ── 엑셀 파싱 ───────────────────────────────────────────
    def parse_excel(self, content: bytes) -> list[dict]:
        """엑셀 바이트를 수강 행 목록으로 변환. 저장은 하지 않는다(미리보기용)."""
        try:
            import openpyxl
        except ImportError:                                  # pragma: no cover
            raise ValueError("서버에 openpyxl이 설치되어 있지 않습니다.")

        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception:
            raise ValueError("엑셀 파일을 열 수 없습니다. .xlsx 파일인지 확인해 주세요.")

        ws = wb.worksheets[0]
        header = [_s(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
        missing = [h for h in EXPECTED_HEADER if h not in header]
        if missing:
            raise ValueError(
                "형식이 다릅니다. 다음 열이 없습니다: " + ", ".join(missing)
                + " · 포털의 '전체 기이수성적' 파일을 그대로 올려주세요.")
        idx = {h: header.index(h) for h in header if h}

        def cell(row, name):
            i = idx.get(name)
            return ws.cell(row, i + 1).value if i is not None else None

        rows: list[dict] = []
        for r in range(2, ws.max_row + 1):
            code = _s(cell(r, "과목번호"))
            name = _s(cell(r, "과목명"))
            if not code or not name:
                continue                                      # 빈 줄·합계 줄 방어
            grade = _s(cell(r, "등급")).upper()
            rows.append({
                "year": int(_num(cell(r, "년도")) or 0) or None,
                "semester": _s(cell(r, "학기")) or None,
                "course_code": code,
                "course_name": name,
                "category": CATEGORY_MAP.get(_s(cell(r, "이수구분")), DEFAULT_CATEGORY),
                "raw_category": _s(cell(r, "이수구분")),
                "credits": _num(cell(r, "학점")) or 0.0,
                "grade": grade or None,
                "grade_point": _num(cell(r, "평점")),
                "retake_of": _s(cell(r, "재수강(대체)과목")) or None,
                "is_passed": grade not in FAIL_GRADES,
            })
        if not rows:
            raise ValueError("읽을 수 있는 수강 기록이 없습니다.")
        return rows

    async def preview_excel(self, db: AsyncSession, student_id: int,
                            content: bytes) -> dict:
        """파싱 + 중복 판정. 저장 전에 학생이 확인할 요약을 만든다."""
        rows = self.parse_excel(content)
        existing = {
            (r[0], r[1], r[2])
            for r in (await db.execute(
                select(StudentCourse.course_code, StudentCourse.year, StudentCourse.semester)
                .where(StudentCourse.student_id == student_id))).all()
        }
        new_count = sum(
            1 for r in rows
            if (r["course_code"], r["year"], r["semester"]) not in existing)

        graded = [(r["credits"], r["grade_point"]) for r in rows if r["grade_point"] is not None]
        cred = sum(c for c, _ in graded)
        return {
            "rows": rows,
            "total": len(rows),
            "new_count": new_count,
            "duplicate_count": len(rows) - new_count,
            "total_credits": sum(r["credits"] for r in rows if r["is_passed"]),
            "gpa_preview": round(sum(c * g for c, g in graded) / cred, 2) if cred else None,
        }

    # ── 조회 ────────────────────────────────────────────────
    async def list_courses(self, db: AsyncSession, student_id: int) -> list[dict]:
        rows = (await db.execute(
            select(StudentCourse, Course)
            .outerjoin(Course, Course.code == StudentCourse.course_code)
            .where(StudentCourse.student_id == student_id)
            .order_by(StudentCourse.year.desc().nullslast(),
                      StudentCourse.semester.desc().nullslast(),
                      StudentCourse.id)
        )).all()
        return [{
            "id": sc.id,
            "year": sc.year,
            "semester": sc.semester,
            "course_code": sc.course_code,
            "course_name": c.name if c else sc.course_code,
            "category": c.category if c else None,
            "credits": float(c.credits) if c and c.credits is not None else 0.0,
            "grade": sc.grade,
            "grade_point": sc.grade_point,
            "is_passed": bool(sc.is_passed),
        } for sc, c in rows]

    # ── 병합 저장 ───────────────────────────────────────────
    async def merge_courses(self, db: AsyncSession, student_id: int,
                            rows: list[dict]) -> dict:
        """엑셀 파싱 결과를 병합한다. 이미 있는 (과목번호, 년도, 학기)는 건드리지 않는다."""
        existing = {
            (r[0], r[1], r[2])
            for r in (await db.execute(
                select(StudentCourse.course_code, StudentCourse.year, StudentCourse.semester)
                .where(StudentCourse.student_id == student_id))).all()
        }
        added, skipped, created_courses = 0, 0, 0

        for row in rows:
            key = (row["course_code"], row.get("year"), row.get("semester"))
            if key in existing:
                skipped += 1
                continue
            if await self._ensure_course(db, row):
                created_courses += 1
            db.add(StudentCourse(
                student_id=student_id,
                course_code=row["course_code"],
                is_passed=row.get("is_passed", True),
                year=row.get("year"),
                semester=row.get("semester"),
                grade=row.get("grade"),
                grade_point=row.get("grade_point"),
                retake_of=row.get("retake_of"),
            ))
            existing.add(key)
            added += 1

        await db.commit()
        gpa = await self.recalc_gpa(db, student_id)
        return {"added": added, "skipped": skipped,
                "created_courses": created_courses, "gpa": gpa}

    async def add_course(self, db: AsyncSession, student_id: int, row: dict) -> dict:
        """수동 1건 추가. 엑셀에 없는 과목(교환학생·인정과목 등)을 직접 넣을 때 쓴다."""
        if not _s(row.get("course_code")) or not _s(row.get("course_name")):
            raise ValueError("과목번호와 과목명은 필수입니다.")
        dup = (await db.execute(
            select(StudentCourse.id).where(
                StudentCourse.student_id == student_id,
                StudentCourse.course_code == row["course_code"],
                StudentCourse.year == row.get("year"),
                StudentCourse.semester == row.get("semester"))
        )).scalar_one_or_none()
        if dup:
            raise ValueError("같은 학기에 이미 등록된 과목입니다.")

        await self._ensure_course(db, row)
        db.add(StudentCourse(
            student_id=student_id,
            course_code=row["course_code"],
            is_passed=row.get("is_passed", True),
            year=row.get("year"),
            semester=row.get("semester"),
            grade=row.get("grade"),
            grade_point=row.get("grade_point"),
        ))
        await db.commit()
        return {"gpa": await self.recalc_gpa(db, student_id)}

    async def delete_course(self, db: AsyncSession, student_id: int, row_id: int) -> dict:
        sc = (await db.execute(
            select(StudentCourse).where(StudentCourse.id == row_id,
                                        StudentCourse.student_id == student_id)
        )).scalar_one_or_none()
        if not sc:
            raise LookupError("해당 수강 기록이 없습니다.")
        await db.delete(sc)
        await db.commit()
        return {"gpa": await self.recalc_gpa(db, student_id)}

    # ── 평점평균 ────────────────────────────────────────────
    async def recalc_gpa(self, db: AsyncSession, student_id: int) -> float | None:
        """학점 가중 평균. P/NP처럼 평점이 없는 과목은 제외한다(포함하면 평균이 왜곡된다)."""
        rows = (await db.execute(
            select(StudentCourse.grade_point, Course.credits)
            .join(Course, Course.code == StudentCourse.course_code)
            .where(StudentCourse.student_id == student_id,
                   StudentCourse.grade_point.isnot(None))
        )).all()
        total_credits = sum(float(c) for _, c in rows if c)
        gpa = None
        if total_credits > 0:
            gpa = round(sum(float(g) * float(c) for g, c in rows if c) / total_credits, 2)
        student = (await db.execute(
            select(Student).where(Student.id == student_id))).scalar_one()
        student.gpa = gpa
        await db.commit()
        return gpa

    # ── 내부 ────────────────────────────────────────────────
    @staticmethod
    async def _ensure_course(db: AsyncSession, row: dict) -> bool:
        """course에 없으면 만든다. 이미 있으면 그대로 쓴다(이름·학점을 덮어쓰지 않는다 —
        관리자가 정리해 둔 마스터 값을 학생 파일이 밀어버리면 안 되므로)."""
        code = row["course_code"]
        found = (await db.execute(
            select(Course).where(Course.code == code))).scalar_one_or_none()
        if found:
            return False
        db.add(Course(
            code=code,
            name=row.get("course_name") or code,
            category=row.get("category") or DEFAULT_CATEGORY,
            credits=float(row.get("credits") or 0),
        ))
        await db.flush()
        return True


student_course_service = StudentCourseService()
